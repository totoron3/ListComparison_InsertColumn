import openpyxl as excel
import numpy as np
import datetime
from spire.xls import Workbook
from spire.xls import FileFormat

print("比較したいリストの一つ目のファイル名を入力してください：")
first_file = input()

Fbook = excel.load_workbook(first_file)
Fsheet = Fbook.active

x = np.array([cell.value for cell in Fsheet['D'][1:]])

x_beside = x.reshape(len(x), 1)

first_col_tmp = np.hstack([x_beside])
first_col_list = first_col_tmp.tolist()

print("比較したいリストの二つ目のファイル名を入力してください：")
second_file = input()

Sbook = excel.load_workbook(second_file)
Ssheet = Sbook.active

x = np.array([cell.value for cell in Ssheet['A'][1:]])
y = np.array([cell.value for cell in Ssheet['B'][1:]])

x_beside = x.reshape(len(x), 1)
y_beside = y.reshape(len(y), 1)

second_col_tmp = np.hstack([x_beside, y_beside])
second_col_list = second_col_tmp.tolist()

tmp_list = []
for i in range(len(first_col_list)):
    for j in range(len(second_col_list)):
        if first_col_list[i][0] == second_col_list[j][1]:
            tmp_list.append([second_col_list[j][0]])
            break

tmp_list = np.array(tmp_list)
tmp_list_beside = tmp_list.reshape(len(tmp_list), 1)
tmp_list_beside = tmp_list_beside.tolist()

Fsheet.insert_cols(3)

Fbook.save(first_file)

workbook = Workbook()

workbook.LoadFromFile(first_file)
ws = workbook.Worksheets.get_Item(0)

i = 2
for array in tmp_list_beside:
    ws.InsertArray(array, i, 3, False)
    i += 1

workbook.SaveToFile(first_file)
workbook.Dispose() 
